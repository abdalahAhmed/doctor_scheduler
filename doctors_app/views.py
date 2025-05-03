from django.shortcuts import render, redirect, get_object_or_404
from .models import Doctor, Clinic, Vacation, Attendance, ScheduleEntry, LocationConfig
from .scheduler import DoctorScheduler, WORK_DAYS, get_upcoming_sunday
from datetime import datetime, timedelta
import pandas as pd
import io
from datetime import date

from django.http import HttpResponse
from collections import defaultdict

def generate_schedule_view(request):
    today = date.today()
    weekday = today.weekday()
    scheduler = DoctorScheduler()

    if request.method == 'POST' or weekday == 1:
        ScheduleEntry.objects.filter(week_start_date=scheduler.week_start).delete()
        scheduler.generate_next_week_schedule()

    return redirect('doctor_dashboard')

def doctors_list(request):
    query = request.GET.get('q', '')
    doctors = Doctor.objects.all()
    existing_clinics = set(Clinic.objects.values_list('name', flat=True))

    for doctor in doctors:
        if doctor.preferred_clinics:
            # 🔥 تنظيف العيادات المفضلة: خلي بس اللي موجود فعلاً
            cleaned_preferred = [name for name in doctor.preferred_clinics if name in existing_clinics]
            if cleaned_preferred != doctor.preferred_clinics:
                doctor.preferred_clinics = cleaned_preferred
                doctor.save(update_fields=['preferred_clinics'])  # 🛠️ نحفظ بس الفيلد المطلوب بدون لمس الباقي

    if query:
        doctors = doctors.filter(
            models.Q(name__icontains=query) | models.Q(specialty__icontains=query)
        )

    return render(request, 'doctors_list.html', {'doctors': doctors})


def clinics_list(request):
    clinics = Clinic.objects.all()
    return render(request, 'clinics_list.html', {'clinics': clinics})

def add_clinic(request):
    if request.method == 'POST':
        name = request.POST['name']
        days = request.POST.getlist('days_available')
        Clinic.objects.create(name=name, days_available=days)
        return redirect('clinics_list')
    return render(request, 'add_clinic.html', {'work_days': WORK_DAYS})

def edit_clinic(request, clinic_id):
    clinic = get_object_or_404(Clinic, pk=clinic_id)
    if request.method == 'POST':
        clinic.name = request.POST['name']
        clinic.days_available = request.POST.getlist('days_available')
        clinic.save()
        return redirect('clinics_list')
    return render(request, 'edit_clinic.html', {
        'clinic': clinic,
        'work_days': WORK_DAYS
    })
def add_doctor(request):
    if request.method == 'POST':
        Doctor.objects.create(
            name=request.POST['name'],
            specialty=request.POST['specialty'],
            preferred_clinics=[s.strip() for s in request.POST['preferred_clinics'].split(',') if s.strip()],
            preferred_days=[s.strip() for s in request.POST['preferred_days'].split(',') if s.strip()],  # ✅ List جاهزة زي ["الأحد-نهاري", "الثلاثاء-ليلي"]
            unavailable_days=[s.strip() for s in request.POST['unavailable_days'].split(',') if s.strip()],  # ✅ List برضه
            shift_preference=[s.strip() for s in request.POST['shift_preference'].split(',') if s.strip()],
            max_sessions=request.POST['max_sessions'],
            has_training='has_training' in request.POST
        )
        return redirect('doctors_list')

    context = {
        'roles': ["رئيس القسم", "نائب رئيس القسم", "أخصائي أول", "استشاري", "أخصائي نساء", "أخصائي رجال", "طبيب مقيم", "طبيب متدرب"],
        'clinic_names': Clinic.objects.values_list('name', flat=True),
        'work_days': WORK_DAYS,
        'shift_types': ['نهاري', 'ليلي']
    }
    return render(request, 'add_doctor.html', context)


def edit_doctor(request, doctor_id):
    doctor = get_object_or_404(Doctor, pk=doctor_id)

    if request.method == 'POST':
        doctor.name = request.POST['name']
        doctor.specialty = request.POST['specialty']
        doctor.preferred_clinics = [s.strip() for s in request.POST['preferred_clinics'].split(',') if s.strip()]
        doctor.preferred_days = [s.strip() for s in request.POST['preferred_days'].split(',') if s.strip()]
        doctor.unavailable_days = [s.strip() for s in request.POST['unavailable_days'].split(',') if s.strip()]
        doctor.shift_preference = [s.strip() for s in request.POST['shift_preference'].split(',') if s.strip()]
        doctor.max_sessions = request.POST['max_sessions']
        doctor.has_training = 'has_training' in request.POST
        doctor.save()
        return redirect('doctors_list')

    context = {
        'doctor': doctor,
        'roles': ["رئيس القسم", "استشاري", "طبيب مقيم", "طبيب متدرب", "أخصائي نساء", "أخصائي رجال"],
        'clinic_names': Clinic.objects.values_list('name', flat=True),
        'work_days': WORK_DAYS,
        'shift_types': ['نهاري', 'ليلي']
    }
    return render(request, 'edit_doctor.html', context)


def vacations_list(request):
    vacations = Vacation.objects.select_related('doctor').all()
    return render(request, 'vacations_list.html', {'vacations': vacations})

def add_vacation(request):
    doctors = Doctor.objects.all()
    if request.method == 'POST':
        Vacation.objects.create(
            doctor_id=request.POST['doctor'],
            start_date=request.POST['start_date'],
            end_date=request.POST['end_date']
        )
        return redirect('vacations_list')
    return render(request, 'add_vacation.html', {'doctors': doctors})

def edit_vacation(request, vacation_id):
    vacation = get_object_or_404(Vacation, pk=vacation_id)
    if request.method == 'POST':
        vacation.start_date = request.POST['start_date']
        vacation.end_date = request.POST['end_date']
        vacation.save()
        return redirect('vacations_list')
    return render(request, 'edit_vacation.html', {'vacation': vacation})

def attendance_list(request):
    attendance_records = Attendance.objects.select_related('doctor').order_by('-date')
    return render(request, 'attendance_list.html', {'attendance_records': attendance_records})

def add_attendance(request):
    doctors = Doctor.objects.all()
    if request.method == 'POST':
        date = request.POST['date']
        for doctor in doctors:
            present_key = f'doctor_{doctor.id}'
            is_present = present_key in request.POST
            Attendance.objects.create(
                doctor=doctor,
                date=date,
                is_present=is_present
            )
        return redirect('attendance_list')
    return render(request, 'add_attendance.html', {'doctors': doctors})

def edit_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, pk=attendance_id)
    if request.method == 'POST':
        attendance.is_present = request.POST['is_present'] == 'true'
        attendance.save()
        return redirect('attendance_list')
    return render(request, 'edit_attendance.html', {'attendance': attendance})

from collections import defaultdict
from .models import ScheduleEntry, LocationConfig
from .scheduler import WORK_DAYS, DoctorScheduler

from .scheduler import WORK_DAYS, get_current_week_start

from .scheduler import WORK_DAYS, get_current_week_start

def current_schedule(request):
    # ✅ العيادات المفعلة
    active_configs = LocationConfig.objects.filter(is_active=True)

    # ✅ بداية الأسبوع الحالي بشكل صحيح
    latest_week_start = get_current_week_start()

    # ✅ سحب الجلسات للأسبوع الحالي
    entries = ScheduleEntry.objects.select_related('doctor', 'clinic')\
        .filter(week_start_date=latest_week_start)\
        .order_by('day', 'session')

    if not entries.exists():
        return render(request, 'current_schedule.html', {
            'days': WORK_DAYS,
            'columns': [],
            'schedule_map': {},
            'doctor_sessions': {},
            'week_start_date': latest_week_start
        })

    # ✅ فلترة اليوم أو النوع لو موجودة
    if request.GET.get('day'):
        entries = entries.filter(day=request.GET['day'])
    if request.GET.get('type') == 'special':
        entries = entries.filter(is_special=True)
    elif request.GET.get('type') == 'coverage':
        entries = entries.filter(is_coverage=True)

    days = WORK_DAYS
    schedule_map = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
    unique_columns = set()

    # ✅ هنا الاختلاف المهم: الأعمدة بناءً على entries مش على العيادات المفعلة
    for entry in entries:
        if entry.clinic:
            session_name = 'نهاري' if entry.session in [1, 2] else 'مسائي'
            col_key = f"{entry.clinic.name} - {session_name}"
            unique_columns.add(col_key)

    # ✅ ملء الجدول
    for entry in entries:
        if entry.clinic:
            session_name = 'نهاري' if entry.session in [1, 2] else 'مسائي'
            period_name = 'الأولى' if entry.session in [1, 3] else 'الثانية'
            col_key = f"{entry.clinic.name} - {session_name}"

            if entry.doctor:
                role_class = entry.doctor.specialty.replace(" ", "_")
                label = f'<span class="doctor-role {role_class}">{entry.doctor.name}</span>'
            else:
                label = '—'

            if entry.is_coverage:
                label += ' <span class="badge bg-warning text-dark">تغطية</span>'
            elif entry.is_special:
                label += ' <span class="badge bg-info text-dark">خاصة</span>'

            schedule_map[entry.day][col_key][period_name] = label

    # ✅ عدد الجلسات لكل طبيب
    scheduler = DoctorScheduler()
    doctor_sessions = scheduler.get_doctor_session_counts(latest_week_start)

    context = {
        'days': days,
        'columns': sorted(unique_columns),
        'schedule_map': dict(schedule_map),
        'doctor_sessions': doctor_sessions,
        'week_start_date': latest_week_start
    }

    return render(request, 'current_schedule.html', context)




import re
from collections import defaultdict
from django.shortcuts import render
from .models import LocationConfig, ScheduleEntry, Doctor
from .scheduler import WORK_DAYS, get_upcoming_sunday, DoctorScheduler

def sort_clinic_key(column_name):
    clinic_name = column_name.split(" - ")[0].strip()

    # (0) العيادات اللي تبدأ برقم
    match = re.match(r'^(\d+)', clinic_name)
    if match:
        return (0, int(match.group()), clinic_name.lower())

    # (1) العيادات اللي تبدأ بـ "Trainer"
    if clinic_name.lower().startswith("trainer"):
        return (1, 0, clinic_name.lower())

    # (2) باقي العيادات
    return (2, 0, clinic_name.lower())

def next_schedule(request):
    active_configs = LocationConfig.objects.filter(is_active=True)
    week_start = get_upcoming_sunday()

    entries = ScheduleEntry.objects.select_related('doctor', 'clinic')\
        .filter(week_start_date=week_start, clinic__in=[cfg.clinic for cfg in active_configs])\
        .order_by('day', 'session')

    if request.GET.get('day'):
        entries = entries.filter(day=request.GET['day'])
    if request.GET.get('type') == 'special':
        entries = entries.filter(is_special=True)
    elif request.GET.get('type') == 'coverage':
        entries = entries.filter(is_coverage=True)

    days = WORK_DAYS
    schedule_map = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
    unique_columns = set()

    for config in active_configs:
        if config.period_1_enabled or config.period_2_enabled:
            unique_columns.add(f"{config.clinic.name} - نهاري")
        if config.period_3_enabled or config.period_4_enabled:
            unique_columns.add(f"{config.clinic.name} - مسائي")

    for entry in entries:
        if entry.clinic:
            session_name = 'نهاري' if entry.session in [1, 2] else 'مسائي'
            period_name = 'الأولى' if entry.session in [1, 3] else 'الثانية'
            col_key = f"{entry.clinic.name} - {session_name}"

            if entry.doctor:
                role_class = entry.doctor.specialty.replace(" ", "_")
                label = f'<span class="doctor-role {role_class}">{entry.doctor.name}</span>'
            else:
                label = '—'

            if entry.is_coverage:
                label += ' <span class="badge bg-warning text-dark">تغطية</span>'
            elif entry.is_special:
                label += ' <span class="badge bg-info text-dark">خاصة</span>'

            schedule_map[entry.day][col_key][period_name] = label

    columns = sorted(unique_columns, key=sort_clinic_key)

    scheduler = DoctorScheduler()
    doctor_sessions = scheduler.get_doctor_session_counts(week_start)

    context = {
        'days': days,
        'columns': columns,
        'schedule_map': dict(schedule_map),
        'doctor_sessions': doctor_sessions,
        'week_start_date': week_start,
        'all_doctors': list(Doctor.objects.values_list('name', flat=True))
    }

    return render(request, 'next_schedule.html', context)



def delete_doctor(request, doctor_id):
    doctor = get_object_or_404(Doctor, pk=doctor_id)
    doctor.delete()
    return redirect('doctors_list')

def delete_clinic(request, clinic_id):
    clinic = get_object_or_404(Clinic, pk=clinic_id)
    clinic.delete()
    return redirect('clinics_list')

def delete_vacation(request, vacation_id):
    vacation = get_object_or_404(Vacation, pk=vacation_id)
    vacation.delete()
    return redirect('vacations_list')

def delete_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, pk=attendance_id)
    attendance.delete()
    return redirect('attendance_list')

def home_view(request):
    return render(request, 'hompage.html')

def dashboard_view(request):
    doctors = Doctor.objects.all()
    clinics = Clinic.objects.all()
    vacations = Vacation.objects.all()
    attendance = Attendance.objects.filter(date=date.today())
    present_count = attendance.filter(is_present=True).count()
    current_schedule = ScheduleEntry.objects.all()

    return render(request, 'dashboard.html', {
        'doctors': doctors,
        'clinics': clinics,
        'vacations': vacations,
        'attendance': attendance,
        'present_count': present_count,
        'current_schedule': current_schedule,
    })

def export_next_schedule_excel(request, lang='ar'):
    from openpyxl import Workbook
    from django.utils import translation
    translation.activate(lang)
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from .scheduler import WORK_DAYS, get_upcoming_sunday
    from .models import LocationConfig, ScheduleEntry
    import io
    from collections import defaultdict

    week_start = get_upcoming_sunday()
    active_configs = LocationConfig.objects.filter(is_active=True)
    entries = ScheduleEntry.objects.select_related('doctor', 'clinic')\
        .filter(week_start_date=week_start, clinic__in=[cfg.clinic for cfg in active_configs])\
        .order_by('day', 'session')

    day_translation = {
        'الأحد': 'Sunday', 'الاثنين': 'Monday', 'الثلاثاء': 'Tuesday',
        'الأربعاء': 'Wednesday', 'الخميس': 'Thursday'
    }
    period_translation = {'الأولى': 'First', 'الثانية': 'Second'}
    shift_translation = {'نهاري': 'Morning', 'مسائي': 'Evening'}

    # تصنيف الأعمدة صباحي ومسائي
    morning_columns = sorted([f"{config.clinic.name} - نهاري" for config in active_configs if config.period_1_enabled or config.period_2_enabled])
    evening_columns = sorted([f"{config.clinic.name} - مسائي" for config in active_configs if config.period_3_enabled or config.period_4_enabled])
    all_columns = morning_columns + evening_columns

    schedule_map = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
    color_map = defaultdict(lambda: '#FFFFFF')

    color_by_specialty = {
        "رئيس القسم": "#444444",
        "نائب رئيس القسم": "#007bff",
        "أخصائي أول": "#6a1b9a",
        "استشاري": "#dc3545",
        "أخصائي نساء": "#f7c6d4",
        "أخصائي رجال": "#28a745",
        "طبيب مقيم": "#17a2b8",
        "طبيب متدرب": "#cfa4ff"
    }

    for entry in entries:
        session_name = 'نهاري' if entry.session in [1, 2] else 'مسائي'
        period = 'الأولى' if entry.session in [1, 3] else 'الثانية'
        col_key = f"{entry.clinic.name} - {session_name}"
        if entry.doctor:
            schedule_map[entry.day][period][col_key] = entry.doctor.name
            specialty = entry.doctor.specialty
            if specialty in color_by_specialty:
                color_map[(entry.day, period, col_key)] = color_by_specialty[specialty]
        else:
            schedule_map[entry.day][period][col_key] = ''

    wb = Workbook()
    ws = wb.active
    ws.title = "Next Week Schedule" if lang == 'en' else "جدول الأسبوع القادم"

    day_col = "Day" if lang == 'en' else "اليوم"
    period_col = "Period" if lang == 'en' else "الفترة"
    translated_columns = []
    for col in all_columns:
        translated_col = col
        if lang == 'en':
            for ar, en in shift_translation.items():
                if ar in translated_col:
                    translated_col = translated_col.replace(ar, en)
        translated_columns.append(translated_col)

    ws.append([day_col, period_col] + translated_columns)

    for day in WORK_DAYS:
        for period in ['الأولى', 'الثانية']:
            display_day = day_translation[day] if lang == 'en' else day
            display_period = period_translation[period] if lang == 'en' else period
            row = [display_day if period == 'الأولى' else '', display_period]
            for col in all_columns:
                row.append(schedule_map[day][period].get(col, ''))
            ws.append(row)

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=1 + len(WORK_DAYS)*2, max_col=2 + len(all_columns)):
        for cell in row:
            cell.alignment = center_align
            if cell.row == 1:
                cell.font = header_font
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ✅ تعديل التلوين على الأعمدة الجديدة all_columns
    for i, day in enumerate(WORK_DAYS):
        for j, period in enumerate(['الأولى', 'الثانية']):
            excel_row = 2 + i * 2 + j
            for k, col in enumerate(all_columns):
                doctor_cell = ws.cell(row=excel_row, column=3 + k)
                color = color_map.get((day, period, col), "#EEEEEE")
                doctor_cell.fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
                doctor_cell.alignment = center_align

    for col in range(1, len(all_columns) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "Next_Week_Schedule.xlsx" if lang == 'en' else "جدول_الأسبوع_القادم.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_current_schedule_excel(request, lang='ar'):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from .scheduler import WORK_DAYS
    from .models import LocationConfig, ScheduleEntry
    import io
    from collections import defaultdict

    latest_week_start = ScheduleEntry.objects.order_by('-week_start_date').values_list('week_start_date', flat=True).first()
    if not latest_week_start:
        return HttpResponse("لا يوجد جدول محفوظ.", content_type='text/plain')

    active_configs = LocationConfig.objects.filter(is_active=True)
    entries = ScheduleEntry.objects.select_related('doctor', 'clinic')\
        .filter(week_start_date=latest_week_start, clinic__in=[cfg.clinic for cfg in active_configs])\
        .order_by('day', 'session')

    day_translation = {
        'الأحد': 'Sunday', 'الاثنين': 'Monday', 'الثلاثاء': 'Tuesday',
        'الأربعاء': 'Wednesday', 'الخميس': 'Thursday'
    }
    period_translation = {'الأولى': 'First', 'الثانية': 'Second'}
    shift_translation = {'نهاري': 'Morning', 'مسائي': 'Evening'}

    # تصنيف الأعمدة صباحي ومسائي
    morning_columns = sorted([f"{config.clinic.name} - نهاري" for config in active_configs if config.period_1_enabled or config.period_2_enabled])
    evening_columns = sorted([f"{config.clinic.name} - مسائي" for config in active_configs if config.period_3_enabled or config.period_4_enabled])
    all_columns = morning_columns + evening_columns

    schedule_map = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
    color_map = defaultdict(lambda: '#FFFFFF')

    color_by_specialty = {
        "رئيس القسم": "#444444",
        "نائب رئيس القسم": "#007bff",
        "أخصائي أول": "#6a1b9a",
        "استشاري": "#dc3545",
        "أخصائي نساء": "#f7c6d4",
        "أخصائي رجال": "#28a745",
        "طبيب مقيم": "#17a2b8",
        "طبيب متدرب": "#cfa4ff"
    }

    for entry in entries:
        session_name = 'نهاري' if entry.session in [1, 2] else 'مسائي'
        period = 'الأولى' if entry.session in [1, 3] else 'الثانية'
        col_key = f"{entry.clinic.name} - {session_name}"
        if entry.doctor:
            schedule_map[entry.day][period][col_key] = entry.doctor.name
            specialty = entry.doctor.specialty
            if specialty in color_by_specialty:
                color_map[(entry.day, period, col_key)] = color_by_specialty[specialty]
        else:
            schedule_map[entry.day][period][col_key] = ''

    wb = Workbook()
    ws = wb.active
    ws.title = "Current Week Schedule" if lang == 'en' else "جدول الأسبوع الحالي"

    day_col = "Day" if lang == 'en' else "اليوم"
    period_col = "Period" if lang == 'en' else "الفترة"
    translated_columns = []
    for col in all_columns:
        translated_col = col
        if lang == 'en':
            for ar, en in shift_translation.items():
                if ar in translated_col:
                    translated_col = translated_col.replace(ar, en)
        translated_columns.append(translated_col)

    ws.append([day_col, period_col] + translated_columns)

    for day in WORK_DAYS:
        for period in ['الأولى', 'الثانية']:
            display_day = day_translation[day] if lang == 'en' else day
            display_period = period_translation[period] if lang == 'en' else period
            row = [display_day if period == 'الأولى' else '', display_period]
            for col in all_columns:
                row.append(schedule_map[day][period].get(col, ''))
            ws.append(row)

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=1 + len(WORK_DAYS)*2, max_col=2 + len(all_columns)):
        for cell in row:
            cell.alignment = center_align
            if cell.row == 1:
                cell.font = header_font
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ✅ تعديل التلوين على الأعمدة الجديدة all_columns
    for i, day in enumerate(WORK_DAYS):
        for j, period in enumerate(['الأولى', 'الثانية']):
            excel_row = 2 + i * 2 + j
            for k, col in enumerate(all_columns):
                doctor_cell = ws.cell(row=excel_row, column=3 + k)
                color = color_map.get((day, period, col), "#EEEEEE")
                doctor_cell.fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
                doctor_cell.alignment = center_align

    for col in range(1, len(all_columns) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "Current_Week_Schedule.xlsx" if lang == 'en' else "جدول_الأسبوع_الحالي.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def save_schedule_changes(request):
    if request.method == 'POST':
        # ✅ تحويل التاريخ من نص إلى كائن تاريخ
        week_start_str = request.POST.get('week_start_date')
        try:
            week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("تاريخ غير صالح", status=400)

        # ✅ حذف الجدول القديم لنفس الأسبوع
        ScheduleEntry.objects.filter(week_start_date=week_start).delete()

        # ✅ حفظ الإدخالات الجديدة
        for key, doctor_name in request.POST.items():
            if key.startswith("assignment__") and doctor_name.strip():
                try:
                    _, day, clinic_shift, period = key.split("__")
                    clinic_name, shift = clinic_shift.strip().split(" - ")
                    session = {
                        ("نهاري", "الأولى"): 1,
                        ("نهاري", "الثانية"): 2,
                        ("مسائي", "الأولى"): 3,
                        ("مسائي", "الثانية"): 4,
                    }.get((shift, period))

                    doctor = Doctor.objects.filter(name=doctor_name).first()
                    clinic = Clinic.objects.filter(name=clinic_name).first()

                    if doctor and clinic and session:
                        ScheduleEntry.objects.create(
                            doctor=doctor,
                            clinic=clinic,
                            day=day,
                            session=session,
                            week_start_date=week_start
                        )
                except Exception as e:
                    print(f"خطأ في الحفظ: {e}")

        return redirect('next_schedule')

    return HttpResponse("يجب استخدام POST.")
    
    
# ✅ نقل جدول الأسبوع القادم إلى الأسبوع الحالي يدويًا
def move_week(request):
    if request.method == 'POST':
        scheduler = DoctorScheduler()
        scheduler.move_next_week_to_current(manual=True)
    return redirect('next_schedule')

# ✅ توليد جدول الأسبوع القادم يدويًا
def generate_schedule_manual(request):
    if request.method == 'POST':
        scheduler = DoctorScheduler()
        scheduler.generate_next_week_schedule(manual=True)
    return redirect('next_schedule')
    
    
def export_current_schedule_en(request):
    return export_current_schedule_excel(request, lang="en")

def export_next_schedule_en(request):
    return export_next_schedule_excel(request, lang="en")
    
    
    
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from .models import LocationConfig

@csrf_exempt
def active_locations_view(request):
    configs = LocationConfig.objects.select_related('clinic').all()

    if request.method == 'POST':
        for config in configs:
            prefix = f"config_{config.id}_"
            config.is_active = prefix + "is_active" in request.POST
            config.period_1_enabled = prefix + "p1" in request.POST
            config.period_2_enabled = prefix + "p2" in request.POST
            config.period_3_enabled = prefix + "p3" in request.POST
            config.period_4_enabled = prefix + "p4" in request.POST
            config.save()
        return redirect('active_locations')  # اسم الـ URL

    return render(request, 'active_locations.html', {'configs': configs})


from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse

@csrf_exempt
def active_locations_editable(request):
    configs = LocationConfig.objects.select_related('clinic').all()

    if request.method == 'POST':
        for config in configs:
            prefix = f"config_{config.id}_"
            config.is_active = bool(request.POST.get(prefix + "is_active"))
            config.period_1_enabled = bool(request.POST.get(prefix + "p1"))
            config.period_2_enabled = bool(request.POST.get(prefix + "p2"))
            config.period_3_enabled = bool(request.POST.get(prefix + "p3"))
            config.period_4_enabled = bool(request.POST.get(prefix + "p4"))
            config.save()
        return redirect('active_locations_editable')

    return render(request, 'active_locations_editable.html', {'configs': configs})
   